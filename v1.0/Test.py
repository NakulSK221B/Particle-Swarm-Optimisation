# -*- coding: utf-8 -*-
"""
Created on Tue Aug 30 11:22:08 2022

@author: NAKUL
"""
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb=load_workbook('Test_WB.xlsx')
ws=wb.active
print(ws['A2'].value)
ws['A2'].value="Change"

print(ws['A2'].value)

wb.save('Test_WB.xlsx')

ws=wb['Sheet2']
wb.create_sheet("New Sheet")

print()
#access value
for row in range(1,11):
    for col in range(1,4):
        Char=get_column_letter(col)
        print(ws[Char+str(row)].value)
        

        
wb.save('Test_WB.xlsx')
