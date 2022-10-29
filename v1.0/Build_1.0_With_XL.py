# -*- coding: utf-8 -*-
"""
Created on Tue Aug 30 18:19:58 2022

@author: NAKUL
"""

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb=load_workbook('D:/BVB/DECENTALISED/PSO/v2.0/Test_WB.xlsx')
# wb=Workbook()
ws=wb['Main']

#Table header Format
ws.merge_cells("A1:J1")
ws['A1'].value="Initial Swarm Generation"
ws.merge_cells("A2:A3")
ws['A2'].value="Particles"
ws.merge_cells("B2:C2")
ws['B2'].value="Positions"
ws['B3'].value="x1"
ws['C3'].value="x2"
ws.merge_cells("D2:E2")
ws['D2'].value="Velocities"
ws['D3'].value="v1"
ws['E3'].value="v2"
ws.merge_cells("F2:F3")
ws['F2'].value="Functional Value"
ws.merge_cells("G2:H2")
ws['G2'].value="Local Best"
ws['G3'].value="P_lb_1"
ws['H3'].value="P_lb_2"
ws.merge_cells("I2:J2")
ws['I2'].value="Global Best"
ws['I3'].value="P_gb_1"
ws['J3'].value="P_gb_2"
no_of_particles=1
Row_no=4
for no_of_particles in range(1,9):
    ws['A'+str(Row_no)].value=no_of_particles
    no_of_particles+=1
    Row_no+=1
print(ws.max_row)
wb.save('D:/BVB/DECENTALISED/PSO/v2.0/Test_WB.xlsx')
# for row in range(1,11):
#     for col in range(1,4):
#         Char=get_column_letter(col)
#         if row==1 and Char=='A':
#             ws.merge_cells(str(Char+str(row))+":"+str(Char+str(row)))
#         ws[Char+str(row)]
#         print(ws[Char+str(row)].value)
        