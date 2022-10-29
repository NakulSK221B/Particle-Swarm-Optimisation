# -*- coding: utf-8 -*-
"""
Created on Thu Sep  8 08:43:10 2022

@author: NAKUL
"""

from openpyxl import load_workbook,drawing
from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.chart import (SurfaceChart,SurfaceChart3D,Reference,Series)
from matplotlib import pyplot as plt
import random
import os

I=50
N=8

Parent_Gen={"Particle No":[1,2,3,4,5,6,7,8],"x1":[2.212,-2.289,-2.393,-0.639,-3.168,0.215,-0.742,-4.563],"x2":[3.009,-2.396,-4.79,1.692,0.706,-2.35,1.934,4.791],
                 "v1":[],
                 "v2":[],
                 "f(x)":[],"lb_1":[],"lb_2":[],"gb_1":[],"gb_2":[]}
Gen={"Particle No":[],"x1":[],"x2":[],"v1":[],"v2":[],"f(x)":[],"lb_1":[],"lb_2":[],"gb_1":[],"gb_2":[]}

func=[]

class pso:
    Gen_Row=1
    wb=load_workbook('Test_WB.xlsx')
    ws_main=wb['Main']
    Test_No=(ws_main['P5'].value)
    ws=wb.create_sheet('Test No.'+str(Test_No+1))
    N=(ws_main.max_row-3)
    def __init__(self): 
        for Gen_Count in range (0,I):
            Current_Gen={"Particle No":[],"x1":[],"x2":[],"v1":[],"v2":[],"f(x)":[],"lb_1":[],"lb_2":[],"gb_1":[],"gb_2":[]}
            if Gen_Count==0:
                pso.Calculate_Func_Value(self, I, Current_Gen)
                
            
    
            
    def Calculate_Func_Value(self,I,Current_Gen):
        if I==0:
            for Func_Calc_Count in range(0,N):
                Func_Value=round(((100*pow((Parent_Gen["x2"][Func_Calc_Count]-pow(Parent_Gen["x1"][Func_Calc_Count],2)),2))+pow((1-(Parent_Gen["x1"][Func_Calc_Count])),2)),3)
                Parent_Gen["f(x)"][Func_Calc_Count].append(Func_Value)
                func.append(Func_Value)
                pso.ws_main['F'+str(Func_Calc_Count+4)]=(Func_Value)
            # print(func)
            Func_Value=0
            Func_Calc_Count=0
            Min_Func=min(Current_Gen["f(x)"])
            Min_Particle=func.index(Min_Func)
            pso.ws_main['F'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="7FFF00")
            func.clear()
        else:
            for Func_Calc_Count in range(8*N,8*(2*N)):
                Func_Value=round(((100*pow((Parent_Gen["x2"][Func_Calc_Count]-pow(Parent_Gen["x1"][Func_Calc_Count],2)),2))+pow((1-(Parent_Gen["x1"][Func_Calc_Count])),2)),3)
                Current_Gen["f(x)"][Func_Calc_Count].append(Func_Value)
                Gen["f(x)"][Func_Calc_Count].append(Func_Value)
                func.append(Func_Value)
            Func_Calc_Count=0
    