# -*- coding: utf-8 -*-
"""
Created on Wed Aug 27 13:01:19 2022

@author: NAKUL
"""

from openpyxl import load_workbook,drawing
from openpyxl.styles import PatternFill,Alignment,Font
from openpyxl.chart import (SurfaceChart,SurfaceChart3D,Reference,Series)
from matplotlib import pyplot as plt
import random
import os
# import sys

# wb=load_workbook('Test_WB.xlsx')
# ws_main=wb['Main']
# Test_No=(ws_main['P5'].value)
# ws=wb.create_sheet('Test No.'+str(Test_No+1))

# N=(ws_main.max_row-3)   #No.Of Particles
Gen_Row_main=1  #Table header Format
I=50 #No.Of Generations

x1=[]
prev_x1=[]
x2=[]
prev_x2=[]
p_lb_1=[]
p_lb_2=[]
p_gb_1=[]
p_gb_2=[]
v1=[]
prev_v1=[]
v2=[]
prev_v2=[]
func=[]
prev_func=[]
plot_x=[]
plot_x_Initial=[]
plot_y=[]
plot_y_Initial=[]
plot_global_x=[]
plot_global_y=[]
# N=8

Min_Particle=0
class pso:
    Gen_Row=1
    wb=load_workbook('Test_WB.xlsx')
    ws_main=wb['Main']
    Test_No=(ws_main['P5'].value)
    ws=wb.create_sheet('Test No.'+str(Test_No+1))
    N=(ws_main.max_row-3)
    def initiate(self,I):
        
        for Gen_Count in range(0,(I+1)):
            if Gen_Count==0:
                pso.Import_Info(Gen_Count)
                pso.Calculate_Func_Value(Gen_Count)
                pso.determine_Local_Best(Gen_Count)
                pso.determine_Global_Best(Gen_Count)
                
            else:
                pso.Create_Gen_Table(Gen_Count)
                pso.Update_Velocity(Gen_Count)
                pso.Update_Position(Gen_Count)
                pso.Calculate_Func_Value(Gen_Count)
                pso.determine_Local_Best(Gen_Count)
                pso.determine_Global_Best(Gen_Count)
                
            # plt.xlabel("X-axis")
            # plt.ylabel("Y-plot")
            # plt.title("Simple x-y plot")
            # plt.scatter(x1, x2, color = "green")
            # plt.pause(0.01)
            # pso.Plot_Excel(N, Gen_Count)
            pso.reset_Local_storage(Gen_Count)
        pso.Plot_Scatter(I)
        pso.Test_No+=1
        pso.ws_main['P5'].value=pso.Test_No
        pso.wb.save('Test_WB.xlsx')
        os.system("start EXCEL.EXE Test_WB.xlsx")
    
    def Plot_Scatter(self,I):
        
        plt.scatter(plot_x_Initial, plot_y_Initial, c ="black",
            linewidths = 2,
            marker ="^",
            edgecolor ="red",
            s = 200)
        # print(plot_x,plot_y)
        plt.scatter(plot_x, plot_y, c ="blue",
            linewidths = 2,
            marker ="s",
            alpha=0.5,
            edgecolor ="green",
            s = 75)
        # plt.scatter(plot_global_x, plot_global_y, c ="red",
        #     linewidths = 2,
        #     marker ="*",
        #     alpha=0.5,
        #     edgecolor ="black",
        #     s = 75)
        plt.title("Particle Swarm Optimisation("+str(I)+" Generations)", loc = 'left')
        plt.xlabel("X1")
        plt.ylabel("X2")
        plt.legend()
        pso.ws.merge_cells('P3:V3')
        pso.ws['P3'].value="Trial No."+str(pso.Test_No+1)
        pso.ws['P3'].alignment=Alignment(horizontal="center",vertical="center")
        pso.ws['P3'].font = Font(name='Calibri',
                    size=25,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        plt.savefig("Plots/Trial No."+str(pso.Test_No)+".png", dpi = 150)
        img = drawing.image.Image("Plots/Trial No."+str(pso.Test_No)+".png")
        img.anchor='M5'
        pso.ws.add_image(img)
        plt.show()
        
    def Calculate_New_Position(self,prev_x1,v1):
        result=round(prev_x1+v1,3)
        if result<-5:
            result=-5
        elif result>5:
            result=5
        else:
            pass
        return result
    
    def Update_Position(self,I):
        for Pos_Particle_count in range(0,pso.N):
            index=(4+Pos_Particle_count+((pso.N+4)*(I-1)))
            x1.append(pso.Calculate_New_Position(prev_x1[Pos_Particle_count], v1[Pos_Particle_count]))
            x2.append(pso.Calculate_New_Position(prev_x2[Pos_Particle_count], v2[Pos_Particle_count]))
            pso.ws['B'+str(index)].value=x1[Pos_Particle_count]
            pso.ws['C'+str(index)].value=x2[Pos_Particle_count]
            Pos_Particle_count+=1
        # print(x1,x2)
                
    
    def Calculate_Velocity(self,v1,x1,p_lb_1,p_gb_1):
        w=0.75
        c1=1.5
        c2=2.0
        r1=round(random.uniform(0.000,1.000),3)
        r2=round(random.uniform(0.000,1.000),3)
        result=((w*v1)+(c1*r1*(p_lb_1-x1))+(c2*r2*(p_gb_1-x1)))
        if result<-5:
            result=-5
        elif result>5:
            result=5
        else:
            pass
        return result
    def Update_Velocity(self,I):
        if I>0:
            # for Vel_Particle_count in range(0,N):
            #For initial swarm
                # x1.append(pso.ws_main['B'+str(Vel_Particle_count+4)].value)
                # x2.append(pso.ws_main['C'+str(Vel_Particle_count+4)].value)
                # func.append(pso.ws_main['F'+str(Vel_Particle_count+4)].value)
                # Vel_Particle_count+=1
            Vel_Particle_count=0
            # print(x1,x2,func)
            for Vel_Particle_count in range(0,pso.N):
                index=(4+Vel_Particle_count+((pso.N+4)*(I-1)))
                pso.ws['D'+str(index)].value=pso.Calculate_Velocity(0, prev_x1[Vel_Particle_count], p_lb_1[Vel_Particle_count], p_gb_1[Vel_Particle_count])
                pso.ws['E'+str(index)].value=pso.Calculate_Velocity(0, prev_x2[Vel_Particle_count], p_lb_2[Vel_Particle_count], p_gb_2[Vel_Particle_count])
                v1.append(pso.ws['D'+str(index)].value)
                v2.append(pso.ws['E'+str(index)].value)
                Vel_Particle_count+=1
            Vel_Particle_count=0
            # print(v1,v2)
    def Import_Info(self,I):
            if I==0:
                for Particle_count in range(1,pso.N+1):
                #For initial swarm
                    # print(type(pso.ws_main['B'+str(Particle_count+3)].value))
                    x1.append(pso.ws_main['B'+str(Particle_count+3)].value)
                    x2.append(pso.ws_main['C'+str(Particle_count+3)].value)
                    plot_x_Initial.append(pso.ws_main['B'+str(Particle_count+3)].value)
                    plot_y_Initial.append(pso.ws_main['C'+str(Particle_count+3)].value)
                    # func.append(pso.ws_main['F'+str(Particle_count+3)].value)
                    Particle_count+=1
            # print(x1)
            # print(x2)
            # print(func)
    def Calculate_Func_Value(self,I):
        if I==0:
            for Func_Calc_Count in range(0,pso.N):
                Func_Value=round(((100*pow((x2[Func_Calc_Count]-pow(x1[Func_Calc_Count],2)),2))+pow((1-(x1[Func_Calc_Count])),2)),3)
                func.append(Func_Value)
                pso.ws_main['F'+str(Func_Calc_Count+4)]=(Func_Value)
            # print(func)
            Func_Value=0
            Func_Calc_Count=0
            Min_Func=min(func)
            Min_Particle=func.index(Min_Func)
            pso.ws_main['F'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="7FFF00")
        else:
            for Func_Calc_Count in range(0,pso.N):
                Func_Value=round(((100*pow((x2[Func_Calc_Count]-pow(x1[Func_Calc_Count],2)),2))+pow((1-(x1[Func_Calc_Count])),2)),3)
                func.append(Func_Value)
            Func_Calc_Count=0
        # print(func)
        
    def determine_Local_Best(self,I):
        if I==0:
            #Local best is equal to current psoition
            for lb_count in range(4,pso.N+4):
                pso.ws_main['G'+str(lb_count)].value=pso.ws_main['B'+str(lb_count)].value
                pso.ws_main['H'+str(lb_count)].value=pso.ws_main['C'+str(lb_count)].value
                p_lb_1.append(pso.ws_main['G'+str(lb_count)].value)
                p_lb_2.append(pso.ws_main['H'+str(lb_count)].value)
                lb_count+=1
            lb_count=0
        else:
            for lb_count in range(0,pso.N):
                index=(4+lb_count+((pso.N+4)*(I-1)))
                func[lb_count]=min(func[lb_count],prev_func[lb_count])
                if func[lb_count]==prev_func[lb_count]:
                    x1[lb_count]=prev_x1[lb_count]
                    x2[lb_count]=prev_x2[lb_count]
                pso.ws['G'+str(index)].value=x1[lb_count]
                pso.ws['H'+str(index)].value=x2[lb_count]
                pso.ws['F'+str(index)].value=func[lb_count]
                lb_count+=1
            Min_Func=min(func)
            Min_Particle=func.index(Min_Func)
            Particle_Index=(4+Min_Particle+((pso.N+4)*(I-1)))
            pso.ws['F'+str(Particle_Index)].fill=PatternFill("solid", start_color="7FFF00")
            
            lb_count=0
            
    def determine_Global_Best(self,I):
        if I==0:
            Min_Func=min(func)
            Min_Particle=func.index(Min_Func)
            plot_global_x.append(pso.ws_main['B'+str(Min_Particle+4)].value)
            plot_global_y.append(pso.ws_main['C'+str(Min_Particle+4)].value)
            for gb_count in range(4,pso.N+4):
                pso.ws_main['I'+str(gb_count)].value=pso.ws_main['B'+str(Min_Particle+4)].value
                pso.ws_main['J'+str(gb_count)].value=pso.ws_main['C'+str(Min_Particle+4)].value
                pso.ws_main['B'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="00FFFF00")
                pso.ws_main['C'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="00FFFF00")
                p_gb_1.append(pso.ws_main['I'+str(gb_count)].value)
                p_gb_2.append(pso.ws_main['J'+str(gb_count)].value)
                gb_count+=1
            gb_count=0
        else:
            Min_Func=min(func)
            Min_Particle=func.index(Min_Func)
            Particle_Index=(4+Min_Particle+((pso.N+4)*(I-1)))
            plot_global_x.append(pso.ws['B'+str(Particle_Index)].value)
            plot_global_y.append(pso.ws['C'+str(Particle_Index)].value)
            pso.ws['B'+str(Particle_Index)].fill= PatternFill("solid", start_color="00FFFF00")
            pso.ws['C'+str(Particle_Index)].fill= PatternFill("solid", start_color="00FFFF00")
            for gb_count in range(0,pso.N):
                index=(4+gb_count+((pso.N+4)*(I-1)))
                pso.ws['I'+str(index)].value=pso.ws['B'+str(Particle_Index)].value
                pso.ws['J'+str(index)].value=pso.ws['C'+str(Particle_Index)].value
        
    def reset_Local_storage(self,I):
        #del previopus values
        prev_func.clear()
        prev_x1.clear()
        prev_x2.clear()
        prev_v1.clear()
        prev_v2.clear()
        #transfer values to prev
        for transfer_count in range(0,len(x1)):
            prev_func.append(func[transfer_count])
            prev_x1.append(x1[transfer_count])
            prev_x2.append(x2[transfer_count])
            if I>0:
                plot_x.append(x1[transfer_count])
                plot_y.append(x2[transfer_count])
            else:
                pass
            # prev_v1.append(v1[transfer_count])
            # prev_v2.append(v2[transfer_count])
            transfer_count+=1
        transfer_count=0
        #Clear Current storage
        func.clear()
        x1.clear()
        x2.clear()
        v1.clear()
        v2.clear()
        # print("After Reset:",prev_func,prev_x1,prev_x2,func,x1,x2,v1,v2)
    def Create_Gen_Table(self,No_of_gen):
        if No_of_gen==0:
            pass
        else:
            pso.ws.merge_cells("A"+str(pso.Gen_Row)+":"+"J"+str(pso.Gen_Row))
            pso.ws['A'+str(pso.Gen_Row)].value="Swarm Generation No:"+str(No_of_gen)
            pso.ws['A'+str(pso.Gen_Row)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws.merge_cells("A"+str(pso.Gen_Row+1)+":"+"A"+str(pso.Gen_Row+2))
            pso.ws['A'+str(pso.Gen_Row+1)].value="Particles"
            pso.ws['A'+str(pso.Gen_Row+1)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws.merge_cells("B"+str(pso.Gen_Row+1)+":"+"C"+str(pso.Gen_Row+1))
            pso.ws['B'+str(pso.Gen_Row+1)].value="Positions"
            pso.ws['B'+str(pso.Gen_Row+1)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['B'+str(pso.Gen_Row+2)].value="x1"
            pso.ws['B'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['C'+str(pso.Gen_Row+2)].value="x2"
            pso.ws['C'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws.merge_cells("D"+str(pso.Gen_Row+1)+":"+"E"+str(pso.Gen_Row+1))
            pso.ws['D'+str(pso.Gen_Row+1)].value="Velocities"
            pso.ws['D'+str(pso.Gen_Row+1)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['D'+str(pso.Gen_Row+2)].value="v1"
            pso.ws['D'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['E'+str(pso.Gen_Row+2)].value="v2"
            pso.ws['E'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws.merge_cells("F"+str(pso.Gen_Row+1)+":"+"F"+str(pso.Gen_Row+2))
            pso.ws['F'+str(pso.Gen_Row+1)].value="Functional Value"
            pso.ws.merge_cells("G"+str(pso.Gen_Row+1)+":"+"H"+str(pso.Gen_Row+1))
            pso.ws['G'+str(pso.Gen_Row+1)].value="Local Best"
            pso.ws['G'+str(pso.Gen_Row+1)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['G'+str(pso.Gen_Row+2)].value="P_lb_1"
            pso.ws['G'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['H'+str(pso.Gen_Row+2)].value="P_lb_2"
            pso.ws['H'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws.merge_cells("I"+str(pso.Gen_Row+1)+":"+"J"+str(pso.Gen_Row+1))
            pso.ws['I'+str(pso.Gen_Row+1)].value="Global Best"
            pso.ws['I'+str(pso.Gen_Row+1)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['I'+str(pso.Gen_Row+2)].value="P_gb_1"
            pso.ws['I'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            pso.ws['J'+str(pso.Gen_Row+2)].value="P_gb_2"
            pso.ws['J'+str(pso.Gen_Row+2)].alignment=Alignment(horizontal="center",vertical="center")
            for no_of_particles in range(1,pso.N+1):
                Row_no=(pso.Gen_Row+2)+no_of_particles
                pso.ws['A'+str(Row_no)].value=no_of_particles
                no_of_particles+=1
        pso.Gen_Row+=(4+pso.N)
        
pso=pso()
pso.initiate(I)
# Test_No+=1
# pso.ws_main['P5'].value=Test_No
# wb.save('Test_WB.xlsx')
