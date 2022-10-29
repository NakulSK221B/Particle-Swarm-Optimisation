# -*- coding: utf-8 -*-
"""
Created on Wed Aug 31 12:06:13 2022

@author: NAKUL
"""

from openpyxl import Workbook,load_workbook
# from openpyxl.utils import get_column_letter
# from datetime import datetime as time

wb=load_workbook('Test_WB.xlsx')
ws_main=wb['Main']
Test_No=(ws_main['P5'].value)
# wb=Workbook()
ws=wb.create_sheet('Test No.'+str(Test_No))
# ws=wb.active
#Table header Format
Gen_Row_main=1
# I=3
# N=8
N=(ws_main.max_row-3)
class pso:
    Gen_Row=1
    def Create_Gen_Table(self,No_of_gen):
        if No_of_gen==0:
            pass
        else:
            ws.merge_cells("A"+str(pso.Gen_Row)+":"+"J"+str(pso.Gen_Row))
            ws['A'+str(pso.Gen_Row)].value="Swarm Generation No:"+str(No_of_gen)
            ws.merge_cells("A"+str(pso.Gen_Row+1)+":"+"A"+str(pso.Gen_Row+2))
            ws['A'+str(pso.Gen_Row+1)].value="Particles"
            ws.merge_cells("B"+str(pso.Gen_Row+1)+":"+"C"+str(pso.Gen_Row+1))
            ws['B'+str(pso.Gen_Row+1)].value="psoitions"
            ws['B'+str(pso.Gen_Row+2)].value="x1"
            ws['C'+str(pso.Gen_Row+2)].value="x2"
            ws.merge_cells("D"+str(pso.Gen_Row+1)+":"+"E"+str(pso.Gen_Row+1))
            ws['D'+str(pso.Gen_Row+1)].value="Velocities"
            ws['D'+str(pso.Gen_Row+2)].value="v1"
            ws['E'+str(pso.Gen_Row+2)].value="v2"
            ws.merge_cells("F"+str(pso.Gen_Row+1)+":"+"F"+str(pso.Gen_Row+2))
            ws['F'+str(pso.Gen_Row+1)].value="Functional Value"
            ws.merge_cells("G"+str(pso.Gen_Row+1)+":"+"H"+str(pso.Gen_Row+1))
            ws['G'+str(pso.Gen_Row+1)].value="Local Best"
            ws['G'+str(pso.Gen_Row+2)].value="P_lb_1"
            ws['H'+str(pso.Gen_Row+2)].value="P_lb_2"
            ws.merge_cells("I"+str(pso.Gen_Row+1)+":"+"J"+str(pso.Gen_Row+1))
            ws['I'+str(pso.Gen_Row+1)].value="Global Best"
            ws['I'+str(pso.Gen_Row+2)].value="P_gb_1"
            ws['J'+str(pso.Gen_Row+2)].value="P_gb_2"
            for no_of_particles in range(1,N+1):
                Row_no=(pso.Gen_Row+2)+no_of_particles
                ws['A'+str(Row_no)].value=no_of_particles
                no_of_particles+=1
        pso.Gen_Row+=(4+N)
        
pso=pso()
pso.Create_Gen_Table(1)
pso.Create_Gen_Table(2)
# Test_No+=1
# ws_main['P5'].value=Test_No
wb.save('Test_WB.xlsx')