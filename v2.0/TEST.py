# -*- coding: utf-8 -*-
"""
Created on Tue Aug 30 21:23:47 2022

@author: NAKUL
"""

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from datetime import datetime as time

Total_Generations=3
Generation_Count=0
x1=[]
prev_x1=[]
x2=[]
prev_x2=[]
v1=[]
v2=[]
func=[]
prev_func=[]
wb=load_workbook('D:/BVB/DECENTALISED/PSO/v2.0/Test_WB.xlsx')
ws_main=wb['Main']
N=(ws_main.max_row-3)
# Test_No=(ws_main['D13'].value)
# ws=wb.create_sheet('Test No.'+str(Test_No))
class pso:
    def Import_Info(self,I):
            if I==0:
                for Particle_count in range(1,N+1):
                #For initial swarm
                    # print(type(ws_main['B'+str(Particle_count+3)].value))
                    x1.append(ws_main['B'+str(Particle_count+3)].value)
                    x2.append(ws_main['C'+str(Particle_count+3)].value)
                    # func.append(ws_main['F'+str(Particle_count+3)].value)
                    Particle_count+=1
            else:
                for Particle_count in range(0,N):
                    # x1.append(ws['B'+str(((12*I)-8)+Particle_count)].value)
                    # x2.append(ws['C'+str(((12*I)-8)+Particle_count)].value)
                    # func.append(ws['F'+str(((12*I)-8)+Particle_count)].value)
                    Particle_count+=1
            print(x1)
            # x1.clear()
            print(x2)
            print(func)
    def Calculate_Func_Value_for_initial_swarm(self):
        for Func_Calc_Count in range(0,N):
            Func_Value=round(((100*pow((x2[Func_Calc_Count]-pow(x1[Func_Calc_Count],2)),2))+pow((1-(x1[Func_Calc_Count])),2)),3)
            func.append(Func_Value)
            ws_main['F'+str(Func_Calc_Count+4)]=(Func_Value)
        print(func)
        Func_Value=0
        Min_Func=min(func)
        Min_Particle=func.index(Min_Func)
        ws_main['F'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="7FFF00")
    # def Calculate_Velocities
    def determine_Local_Best(self,I):
        if I==0:
            #Local best is equal to current psoition
            for lb_count in range(4,N+4):
                ws_main['G'+str(lb_count)].value=ws_main['B'+str(lb_count)].value
                ws_main['H'+str(lb_count)].value=ws_main['C'+str(lb_count)].value
                lb_count+=1
        lb_count=0
        # else:
            
    def determine_Global_Best(self,I):
        if I==0:
            for gb_count in range(4,N+4):
                Min_Func=min(func)
                Min_Particle=func.index(Min_Func)
                ws_main['I'+str(gb_count)].value=ws_main['B'+str(Min_Particle+4)].value
                ws_main['J'+str(gb_count)].value=ws_main['C'+str(Min_Particle+4)].value
                ws_main['B'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="00FFFF00")
                ws_main['C'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="00FFFF00")
                gb_count+=1
        gb_count=0



pso=pso()
pso.Import_Info(0)
pso.Calculate_Func_Value_for_initial_swarm()
pso.determine_Local_Best(0)
pso.determine_Global_Best(0)
wb.save('D:/BVB/DECENTALISED/PSO/v2.0/Test_WB.xlsx')