# -*- coding: utf-8 -*-
"""
Created on Wed Aug 31 13:01:19 2022

@author: NAKUL
"""
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
import random
wb=load_workbook('Test_WB.xlsx')
ws_main=wb['Main']
Test_No=(ws_main['P5'].value)
ws=wb.create_sheet('Test No.'+str(Test_No+1))
#Table header Format
Gen_Row_main=1
I=2
x1=[]
prev_x1=[]
x2=[]
prev_x2=[]
p_lb_1=[]
p_lb_2=[]
p_gb_1=[]
p_gb_2=[]
v1=[]
prev_v1=[]
v2=[]
prev_v2=[]
func=[]
prev_func=[]
# N=8
N=(ws_main.max_row-3)
Min_Particle=0
class pso:
    Gen_Row=1
    def initiate(self,I,N):
        for Gen_Count in range(0,(I+1)):
            if Gen_Count==0:
                pso.Import_Info(Gen_Count)
                pso.Calculate_Func_Value(Gen_Count)
                pso.determine_Local_Best(Gen_Count)
                pso.determine_Global_Best(Gen_Count)
                pso.reset_Local_storage()
            else:
                pso.Create_Gen_Table(Gen_Count)
                pso.Update_Velocity(Gen_Count)
                
    def Calculate_New_Position(self,x1,v1)
    
    def Calculate_Velocity(self,v1,x1,p_lb_1,p_gb_1):
        w=0.75
        c1=1.5
        c2=2.0
        r1=round(random.uniform(0.000,1.000),3)
        r2=round(random.uniform(0.000,1.000),3)
        result=((w*v1)+(c1*r1*(p_lb_1-x1))+(c2*r2*(p_gb_1-x1)))
        if result<-5:
            result=-5
        elif result>5:
            result=5
        else:
            pass
        return result
    def Update_Velocity(self,I):
        if I==1:
            for Vel_Particle_count in range(0,N):
            #For initial swarm
                x1.append(ws_main['B'+str(Vel_Particle_count+4)].value)
                x2.append(ws_main['C'+str(Vel_Particle_count+4)].value)
                func.append(ws_main['F'+str(Vel_Particle_count+4)].value)
                Vel_Particle_count+=1
            Vel_Particle_count=0
            # print(x1,x2,func)
            for Vel_Particle_count in range(0,N):
                index=(4+Vel_Particle_count+((N+4)*(I-1)))
                ws['D'+str(index)].value=pso.Calculate_Velocity(0, x1[Vel_Particle_count], p_lb_1[Vel_Particle_count], p_gb_1[Vel_Particle_count])
                ws['E'+str(index)].value=pso.Calculate_Velocity(0, x2[Vel_Particle_count], p_lb_2[Vel_Particle_count], p_gb_2[Vel_Particle_count])
                v1.append(ws['D'+str(index)].value)
                v2.append(ws['E'+str(index)].value)
                Vel_Particle_count+=1
            Vel_Particle_count=0
            # print(v1,v2)
    def Import_Info(self,I):
            if I==0:
                for Particle_count in range(1,N+1):
                #For initial swarm
                    # print(type(ws_main['B'+str(Particle_count+3)].value))
                    x1.append(ws_main['B'+str(Particle_count+3)].value)
                    x2.append(ws_main['C'+str(Particle_count+3)].value)
                    # func.append(ws_main['F'+str(Particle_count+3)].value)
                    Particle_count+=1
            # print(x1)
            # print(x2)
            # print(func)
    def Calculate_Func_Value(self,I):
        if I==0:
            for Func_Calc_Count in range(0,N):
                Func_Value=round(((100*pow((x2[Func_Calc_Count]-pow(x1[Func_Calc_Count],2)),2))+pow((1-(x1[Func_Calc_Count])),2)),3)
                func.append(Func_Value)
                ws_main['F'+str(Func_Calc_Count+4)]=(Func_Value)
            # print(func)
            Func_Value=0
            Min_Func=min(func)
            Min_Particle=func.index(Min_Func)
            ws_main['F'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="7FFF00")
        # else:
    
    def determine_Local_Best(self,I):
        if I==0:
            #Local best is equal to current psoition
            for lb_count in range(4,N+4):
                ws_main['G'+str(lb_count)].value=ws_main['B'+str(lb_count)].value
                ws_main['H'+str(lb_count)].value=ws_main['C'+str(lb_count)].value
                p_lb_1.append(ws_main['G'+str(lb_count)].value)
                p_lb_2.append(ws_main['H'+str(lb_count)].value)
                lb_count+=1
        lb_count=0
        # else:
        # else:
            
    def determine_Global_Best(self,I):
        if I==0:
            for gb_count in range(4,N+4):
                Min_Func=min(func)
                Min_Particle=func.index(Min_Func)
                ws_main['I'+str(gb_count)].value=ws_main['B'+str(Min_Particle+4)].value
                ws_main['J'+str(gb_count)].value=ws_main['C'+str(Min_Particle+4)].value
                ws_main['B'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="00FFFF00")
                ws_main['C'+str(Min_Particle+4)].fill= PatternFill("solid", start_color="00FFFF00")
                p_gb_1.append(ws_main['I'+str(gb_count)].value)
                p_gb_2.append(ws_main['J'+str(gb_count)].value)
                gb_count+=1
        gb_count=0
           
    def reset_Local_storage(self):
        #del previopus values
        prev_func.clear()
        prev_x1.clear()
        prev_x2.clear()
        prev_v1.clear()
        prev_v2.clear()
        #transfer values to prev
        for transfer_count in range(0,len(x1)):
            prev_func.append(func[transfer_count])
            prev_x1.append(x1[transfer_count])
            prev_x2.append(x2[transfer_count])
            # prev_v1.append(v1[transfer_count])
            # prev_v2.append(v2[transfer_count])
            transfer_count+=1
        transfer_count=0
        #Clear Current storage
        func.clear()
        x1.clear()
        x2.clear()
        v1.clear()
        v2.clear()
        # print(prev_func,prev_x1,prev_x2,func,x1,x2,v1,v2)
    def Create_Gen_Table(self,No_of_gen):
        if No_of_gen==0:
            pass
        else:
            ws.merge_cells("A"+str(pso.Gen_Row)+":"+"J"+str(pso.Gen_Row))
            ws['A'+str(pso.Gen_Row)].value="Swarm Generation No:"+str(No_of_gen)
            ws.merge_cells("A"+str(pso.Gen_Row+1)+":"+"A"+str(pso.Gen_Row+2))
            ws['A'+str(pso.Gen_Row+1)].value="Particles"
            ws.merge_cells("B"+str(pso.Gen_Row+1)+":"+"C"+str(pso.Gen_Row+1))
            ws['B'+str(pso.Gen_Row+1)].value="psoitions"
            ws['B'+str(pso.Gen_Row+2)].value="x1"
            ws['C'+str(pso.Gen_Row+2)].value="x2"
            ws.merge_cells("D"+str(pso.Gen_Row+1)+":"+"E"+str(pso.Gen_Row+1))
            ws['D'+str(pso.Gen_Row+1)].value="Velocities"
            ws['D'+str(pso.Gen_Row+2)].value="v1"
            ws['E'+str(pso.Gen_Row+2)].value="v2"
            ws.merge_cells("F"+str(pso.Gen_Row+1)+":"+"F"+str(pso.Gen_Row+2))
            ws['F'+str(pso.Gen_Row+1)].value="Functional Value"
            ws.merge_cells("G"+str(pso.Gen_Row+1)+":"+"H"+str(pso.Gen_Row+1))
            ws['G'+str(pso.Gen_Row+1)].value="Local Best"
            ws['G'+str(pso.Gen_Row+2)].value="P_lb_1"
            ws['H'+str(pso.Gen_Row+2)].value="P_lb_2"
            ws.merge_cells("I"+str(pso.Gen_Row+1)+":"+"J"+str(pso.Gen_Row+1))
            ws['I'+str(pso.Gen_Row+1)].value="Global Best"
            ws['I'+str(pso.Gen_Row+2)].value="P_gb_1"
            ws['J'+str(pso.Gen_Row+2)].value="P_gb_2"
            for no_of_particles in range(1,N+1):
                Row_no=(pso.Gen_Row+2)+no_of_particles
                ws['A'+str(Row_no)].value=no_of_particles
                no_of_particles+=1
        pso.Gen_Row+=(4+N)
        
pso=pso()
pso.initiate(I, N)
Test_No+=1
ws_main['P5'].value=Test_No
wb.save('Test_WB.xlsx')